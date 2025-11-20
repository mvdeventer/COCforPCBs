import openpyxl
import pandas as pd

# Load workbook
wb = openpyxl.load_workbook("BT3413A-8 (Bill of Materials).xlsx")
print("Sheet names:", wb.sheetnames)

ws = wb.active
print("\nFirst 15 rows (raw):")
for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
    print(f"Row {i}: {row}")

# Also check with pandas
df = pd.read_excel("BT3413A-8 (Bill of Materials).xlsx")
print("\nPandas DataFrame Info:")
print(f"Columns: {df.columns.tolist()}")
print(f"\nFirst row values:")
print(df.iloc[0])
